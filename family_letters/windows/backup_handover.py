import os
import sys
import shutil
from datetime import datetime

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QPushButton,
    QTableWidget, QTableWidgetItem, QTextEdit, QFileDialog,
    QHeaderView, QLabel, QProgressBar, QMessageBox, QAbstractItemView,
    QComboBox, QCheckBox,
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QPainter, QBrush, QPen

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from database import execute_query, execute_query_returning, execute_update, get_statistics, DB_PATH

try:
    import openpyxl
    from openpyxl.styles import Font as XlFont, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image as RLImage, PageBreak
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

STYLESHEET = """
QWidget {
    background-color: #f5f7fa;
    font-family: "Microsoft YaHei", "SimHei", sans-serif;
    font-size: 13px;
}
QGroupBox {
    font-weight: bold;
    font-size: 14px;
    border: 1px solid #d0d7de;
    border-radius: 8px;
    margin-top: 12px;
    padding: 16px 12px 12px 12px;
    background-color: #ffffff;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 16px;
    padding: 0 6px;
    color: #1a73e8;
}
QPushButton {
    background-color: #1a73e8;
    color: #ffffff;
    border: none;
    border-radius: 6px;
    padding: 8px 20px;
    font-size: 13px;
    min-height: 32px;
}
QPushButton:hover {
    background-color: #1557b0;
}
QPushButton:pressed {
    background-color: #0d47a1;
}
QPushButton:disabled {
    background-color: #b0bec5;
}
QTableWidget {
    background-color: #ffffff;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    gridline-color: #e8ecf0;
    selection-background-color: #e3f2fd;
    selection-color: #1a1a1a;
}
QTableWidget::item {
    padding: 4px 8px;
}
QHeaderView::section {
    background-color: #e8ecf0;
    color: #333;
    font-weight: bold;
    padding: 6px 8px;
    border: none;
    border-bottom: 2px solid #1a73e8;
}
QTextEdit {
    background-color: #1e1e2e;
    color: #a6e3a1;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    padding: 8px;
    font-family: "Consolas", "Microsoft YaHei", monospace;
    font-size: 12px;
}
QLabel {
    color: #333;
}
"""


class BarChartWidget(QWidget):
    def __init__(self, data=None, title="", parent=None):
        super().__init__(parent)
        self._data = data or {}
        self._title = title
        self.setMinimumHeight(180)
        self.setMaximumHeight(220)

    def set_data(self, data, title=""):
        self._data = data
        self._title = title
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        w, h = self.width(), self.height()
        margin_left, margin_right, margin_top, margin_bottom = 60, 20, 30, 30
        chart_w = w - margin_left - margin_right
        chart_h = h - margin_top - margin_bottom

        if not self._data:
            painter.setPen(QColor("#999"))
            painter.setFont(QFont("Microsoft YaHei", 11))
            painter.drawText(self.rect(), Qt.AlignCenter, "暂无数据")
            painter.end()
            return

        painter.setPen(QColor("#1a73e8"))
        painter.setFont(QFont("Microsoft YaHei", 10, QFont.Bold))
        painter.drawText(0, 0, w, margin_top, Qt.AlignCenter, self._title)

        max_val = max(self._data.values()) if self._data else 1
        if max_val == 0:
            max_val = 1
        bar_count = len(self._data)
        bar_gap = max(chart_w // (bar_count * 3), 4)
        bar_w = (chart_w - bar_gap * (bar_count + 1)) // max(bar_count, 1)
        bar_w = max(bar_w, 12)

        painter.setPen(QPen(QColor("#d0d7de"), 1))
        painter.drawLine(margin_left, margin_top, margin_left, h - margin_bottom)
        painter.drawLine(margin_left, h - margin_bottom, w - margin_right, h - margin_bottom)

        colors = [
            "#1a73e8", "#34a853", "#fbbc04", "#ea4335",
            "#9c27b0", "#00bcd4", "#ff9800", "#795548",
        ]
        painter.setFont(QFont("Microsoft YaHei", 8))
        for i, (label, value) in enumerate(self._data.items()):
            bar_h = int((value / max_val) * chart_h)
            x = margin_left + bar_gap + i * (bar_w + bar_gap)
            y = h - margin_bottom - bar_h
            color = QColor(colors[i % len(colors)])
            painter.setBrush(QBrush(color))
            painter.setPen(Qt.NoPen)
            painter.drawRoundedRect(x, y, bar_w, bar_h, 3, 3)

            painter.setPen(QColor("#333"))
            display_label = label if len(label) <= 6 else label[:5] + "…"
            painter.drawText(x - 4, h - margin_bottom + 4, bar_w + 8, 20,
                             Qt.AlignCenter, display_label)
            painter.setPen(QColor("#666"))
            painter.drawText(x, y - 16, bar_w, 16, Qt.AlignCenter, str(value))

        painter.end()


class BackupHandoverWindow(QWidget):
    data_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("备份交接")
        self.setMinimumSize(960, 780)
        self.setStyleSheet(STYLESHEET)
        self._setup_ui()
        self._refresh_backup_list()
        self._refresh_handover_table()
        self._refresh_statistics()
        self._log("系统就绪，备份交接窗口已加载。")

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)

        layout.addWidget(self._create_backup_group())
        layout.addWidget(self._create_export_group())
        layout.addWidget(self._create_handover_group())
        layout.addWidget(self._create_statistics_group())

        log_label = QLabel("操作日志")
        log_label.setStyleSheet("font-weight:bold; font-size:13px; color:#555; margin-top:4px;")
        layout.addWidget(log_label)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMaximumHeight(120)
        layout.addWidget(self.log_text)

    def _create_backup_group(self):
        group = QGroupBox("数据备份")
        layout = QVBoxLayout(group)

        btn_row = QHBoxLayout()
        self.btn_backup = QPushButton("备份数据库")
        self.btn_backup.clicked.connect(self._backup_database)
        self.btn_restore = QPushButton("恢复数据库")
        self.btn_restore.clicked.connect(self._restore_database)
        self.btn_restore.setStyleSheet(
            "QPushButton{background-color:#ea4335;} QPushButton:hover{background-color:#c5221f;}"
        )
        btn_row.addWidget(self.btn_backup)
        btn_row.addWidget(self.btn_restore)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        self.backup_table = QTableWidget(0, 4)
        self.backup_table.setHorizontalHeaderLabels(["备份时间", "文件路径", "大小", "说明"])
        self.backup_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.backup_table.horizontalHeader().setStretchLastSection(True)
        self.backup_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.backup_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.backup_table.setMaximumHeight(160)
        layout.addWidget(self.backup_table)

        return group

    def _create_export_group(self):
        group = QGroupBox("导出打印册")
        layout = QHBoxLayout(group)

        self.btn_export_excel = QPushButton("导出为Excel")
        self.btn_export_excel.clicked.connect(self._export_excel)
        self.btn_export_pdf = QPushButton("导出为PDF")
        self.btn_export_pdf.clicked.connect(self._export_pdf)
        self.btn_export_html = QPushButton("导出为HTML")
        self.btn_export_html.clicked.connect(self._export_html)

        layout.addWidget(self.btn_export_excel)
        layout.addWidget(self.btn_export_pdf)
        layout.addWidget(self.btn_export_html)
        layout.addStretch()
        return group

    def _create_handover_group(self):
        group = QGroupBox("移交清单")
        layout = QVBoxLayout(group)

        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("借阅状态："))
        self.filter_borrow_combo = QComboBox()
        self.filter_borrow_combo.addItems(["全部", "借出", "已归还", "无"])
        self.filter_borrow_combo.setMinimumWidth(90)
        self.filter_borrow_combo.currentTextChanged.connect(self._refresh_handover_table)
        filter_row.addWidget(self.filter_borrow_combo)

        filter_row.addWidget(QLabel("修复状态："))
        self.filter_restoration_combo = QComboBox()
        self.filter_restoration_combo.addItems(["全部", "良好", "轻微损毁", "需修复", "已修复"])
        self.filter_restoration_combo.setMinimumWidth(90)
        self.filter_restoration_combo.currentTextChanged.connect(self._refresh_handover_table)
        filter_row.addWidget(self.filter_restoration_combo)

        self.filter_private_check = QCheckBox("仅私密")
        self.filter_private_check.stateChanged.connect(self._refresh_handover_table)
        filter_row.addWidget(self.filter_private_check)
        filter_row.addStretch()
        layout.addLayout(filter_row)

        self.handover_table = QTableWidget(0, 6)
        self.handover_table.setHorizontalHeaderLabels(
            ["编号", "信件标题", "寄信人", "收信人", "保存状态", "借阅状态"]
        )
        self.handover_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.handover_table.horizontalHeader().setStretchLastSection(True)
        self.handover_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.handover_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.handover_table.setMaximumHeight(180)
        layout.addWidget(self.handover_table)

        btn_row = QHBoxLayout()
        self.btn_gen_handover = QPushButton("生成移交清单")
        self.btn_gen_handover.clicked.connect(self._generate_handover)
        self.btn_print_handover = QPushButton("打印清单")
        self.btn_print_handover.clicked.connect(self._print_handover)
        self.btn_print_handover.setStyleSheet(
            "QPushButton{background-color:#34a853;} QPushButton:hover{background-color:#2d8e47;}"
        )
        btn_row.addWidget(self.btn_gen_handover)
        btn_row.addWidget(self.btn_print_handover)
        btn_row.addStretch()
        layout.addLayout(btn_row)

        return group

    def _create_statistics_group(self):
        group = QGroupBox("数据统计")
        layout = QHBoxLayout(group)

        self.chart_category = BarChartWidget(title="按分类统计")
        self.chart_year = BarChartWidget(title="按年份统计")
        self.chart_status = BarChartWidget(title="按保存状态统计")

        layout.addWidget(self.chart_category)
        layout.addWidget(self.chart_year)
        layout.addWidget(self.chart_status)

        return group

    def _log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")
        sb = self.log_text.verticalScrollBar()
        sb.setValue(sb.maximum())

    def _refresh_backup_list(self):
        rows = execute_query_returning(
            "SELECT backup_date, file_path, file_size, description FROM backups ORDER BY backup_date DESC"
        )
        self.backup_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            self.backup_table.setItem(i, 0, QTableWidgetItem(str(row.get("backup_date", ""))))
            self.backup_table.setItem(i, 1, QTableWidgetItem(str(row.get("file_path", ""))))
            self.backup_table.setItem(i, 2, QTableWidgetItem(str(row.get("file_size", ""))))
            self.backup_table.setItem(i, 3, QTableWidgetItem(str(row.get("description", ""))))

    def _refresh_handover_table(self):
        rows = execute_query_returning("""
            SELECT l.id, l.title, p1.name AS sender_name, p2.name AS receiver_name,
                   l.restoration_status, l.is_private,
                   COALESCE(br.status, '无') AS borrow_status
            FROM letters l
            LEFT JOIN people p1 ON l.sender_id = p1.id
            LEFT JOIN people p2 ON l.receiver_id = p2.id
            LEFT JOIN (
                SELECT letter_id, status FROM borrow_records
                WHERE id IN (SELECT MAX(id) FROM borrow_records GROUP BY letter_id)
            ) br ON br.letter_id = l.id
            ORDER BY l.id
        """)

        borrow_filter = self.filter_borrow_combo.currentText()
        if borrow_filter != "全部":
            rows = [r for r in rows if r.get("borrow_status", "无") == borrow_filter]

        restoration_filter = self.filter_restoration_combo.currentText()
        if restoration_filter != "全部":
            rows = [r for r in rows if r.get("restoration_status", "") == restoration_filter]

        if self.filter_private_check.isChecked():
            rows = [r for r in rows if r.get("is_private")]

        self.handover_table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            self.handover_table.setItem(i, 0, QTableWidgetItem(str(row.get("id", ""))))
            self.handover_table.setItem(i, 1, QTableWidgetItem(str(row.get("title", ""))))
            self.handover_table.setItem(i, 2, QTableWidgetItem(str(row.get("sender_name", ""))))
            self.handover_table.setItem(i, 3, QTableWidgetItem(str(row.get("receiver_name", ""))))

            restoration = str(row.get("restoration_status", ""))
            rest_item = QTableWidgetItem(restoration)
            if restoration == "需修复":
                rest_item.setForeground(QColor("#ea4335"))
            elif restoration == "良好":
                rest_item.setForeground(QColor("#34a853"))
            self.handover_table.setItem(i, 4, rest_item)

            borrow = str(row.get("borrow_status", "无"))
            borrow_item = QTableWidgetItem(borrow)
            if borrow == "借出":
                borrow_item.setForeground(QColor("#fbbc04"))
            self.handover_table.setItem(i, 5, borrow_item)

    def _get_handover_filter_desc(self):
        parts = []
        borrow = self.filter_borrow_combo.currentText()
        if borrow != "全部":
            parts.append(f"借阅状态={borrow}")
        restoration = self.filter_restoration_combo.currentText()
        if restoration != "全部":
            parts.append(f"修复状态={restoration}")
        if self.filter_private_check.isChecked():
            parts.append("仅私密")
        return "；".join(parts) if parts else "无筛选"

    def _refresh_statistics(self):
        category_data = {}
        rows = execute_query_returning(
            "SELECT COALESCE(category,'未分类') AS cat, COUNT(*) AS cnt "
            "FROM letters GROUP BY cat ORDER BY cnt DESC LIMIT 10"
        )
        for r in rows:
            category_data[r["cat"]] = r["cnt"]
        self.chart_category.set_data(category_data, "按分类统计")

        year_data = {}
        rows = execute_query_returning(
            "SELECT substr(send_date,1,4) AS yr, COUNT(*) AS cnt "
            "FROM letters WHERE send_date IS NOT NULL AND send_date != '' "
            "GROUP BY yr ORDER BY yr LIMIT 15"
        )
        for r in rows:
            year_data[r["yr"]] = r["cnt"]
        self.chart_year.set_data(year_data, "按年份统计")

        status_data = {}
        rows = execute_query_returning(
            "SELECT COALESCE(restoration_status,'未知') AS st, COUNT(*) AS cnt "
            "FROM letters GROUP BY st ORDER BY cnt DESC"
        )
        for r in rows:
            status_data[r["st"]] = r["cnt"]
        self.chart_status.set_data(status_data, "按保存状态统计")

    def _backup_database(self):
        if not os.path.exists(DB_PATH):
            QMessageBox.warning(self, "错误", "数据库文件不存在！")
            self._log("备份失败：数据库文件不存在。")
            return

        dest_dir = QFileDialog.getExistingDirectory(self, "选择备份保存目录")
        if not dest_dir:
            return

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest_filename = f"family_letters_{ts}.db"
        dest_path = os.path.join(dest_dir, dest_filename)

        try:
            shutil.copy2(DB_PATH, dest_path)
            file_size = os.path.getsize(dest_path)
            size_str = f"{file_size / 1024:.1f} KB" if file_size < 1024 * 1024 else f"{file_size / (1024*1024):.2f} MB"
            execute_query(
                "INSERT INTO backups (file_path, backup_date, description, file_size) VALUES (?, datetime('now','localtime'), ?, ?)",
                (dest_path, f"手动备份 - {ts}", size_str),
            )
            self._refresh_backup_list()
            self._log(f"数据库已备份至：{dest_path}（{size_str}）")
            QMessageBox.information(self, "备份成功", f"数据库已备份至：\n{dest_path}")
        except Exception as e:
            self._log(f"备份失败：{e}")
            QMessageBox.critical(self, "备份失败", str(e))

    def _restore_database(self):
        reply = QMessageBox.warning(
            self, "确认恢复",
            "恢复数据库将覆盖当前数据，此操作不可撤销！\n确定要继续吗？",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        src_path, _ = QFileDialog.getOpenFileName(
            self, "选择数据库备份文件", "", "数据库文件 (*.db)"
        )
        if not src_path:
            return

        try:
            shutil.copy2(src_path, DB_PATH)
            self._refresh_backup_list()
            self._refresh_handover_table()
            self._refresh_statistics()
            self._log(f"数据库已从备份恢复：{src_path}")
            QMessageBox.information(self, "恢复成功", "数据库已成功恢复！")
        except Exception as e:
            self._log(f"恢复失败：{e}")
            QMessageBox.critical(self, "恢复失败", str(e))

    def _get_all_letters(self):
        return execute_query_returning("""
            SELECT l.*, p1.name AS sender_name, p2.name AS receiver_name
            FROM letters l
            LEFT JOIN people p1 ON l.sender_id = p1.id
            LEFT JOIN people p2 ON l.receiver_id = p2.id
            ORDER BY l.send_date, l.id
        """)

    def _get_filtered_handover_letters(self):
        rows = execute_query_returning("""
            SELECT l.*, p1.name AS sender_name, p2.name AS receiver_name,
                   COALESCE(br.status, '无') AS borrow_status
            FROM letters l
            LEFT JOIN people p1 ON l.sender_id = p1.id
            LEFT JOIN people p2 ON l.receiver_id = p2.id
            LEFT JOIN (
                SELECT letter_id, status FROM borrow_records
                WHERE id IN (SELECT MAX(id) FROM borrow_records GROUP BY letter_id)
            ) br ON br.letter_id = l.id
            ORDER BY l.send_date, l.id
        """)

        borrow_filter = self.filter_borrow_combo.currentText()
        if borrow_filter != "全部":
            rows = [r for r in rows if r.get("borrow_status", "无") == borrow_filter]

        restoration_filter = self.filter_restoration_combo.currentText()
        if restoration_filter != "全部":
            rows = [r for r in rows if r.get("restoration_status", "") == restoration_filter]

        if self.filter_private_check.isChecked():
            rows = [r for r in rows if r.get("is_private")]

        return rows

    def _export_excel(self):
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "缺少依赖", "未安装 openpyxl 库，无法导出 Excel。\n请运行：pip install openpyxl")
            self._log("导出Excel失败：缺少 openpyxl 库。")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", "家书集.xlsx", "Excel文件 (*.xlsx)"
        )
        if not path:
            return

        try:
            letters = self._get_filtered_handover_letters()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "家书集"

            filter_desc = self._get_handover_filter_desc()
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            ws.merge_cells("A1:L1")
            ws.cell(row=1, column=1, value="家书集").font = XlFont(
                name="Microsoft YaHei", bold=True, size=16, color="1A73E8"
            )
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            ws.merge_cells("A2:L2")
            ws.cell(row=2, column=1, value=f"生成时间：{now_str}　筛选条件：{filter_desc}　共计 {len(letters)} 条记录").font = XlFont(
                name="Microsoft YaHei", size=11, color="333333"
            )

            headers = ["编号", "标题", "寄信人", "收信人", "寄出日期", "收到日期",
                        "寄出地点", "收到地点", "内容", "分类", "保存状态", "备注"]
            header_font = XlFont(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            content_font = XlFont(name="Microsoft YaHei", size=10)
            content_align = Alignment(vertical="top", wrap_text=True)
            for row_idx, lt in enumerate(letters, 4):
                values = [
                    lt.get("id"), lt.get("title", ""), lt.get("sender_name", ""),
                    lt.get("receiver_name", ""), lt.get("send_date", ""),
                    lt.get("receive_date", ""), lt.get("send_location", ""),
                    lt.get("receive_location", ""), lt.get("content", ""),
                    lt.get("category", ""), lt.get("restoration_status", ""),
                    lt.get("notes", ""),
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = content_font
                    cell.alignment = content_align
                    cell.border = thin_border

            col_widths = [6, 20, 10, 10, 14, 14, 12, 12, 50, 10, 10, 20]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            wb.save(path)
            self._log(f"已导出Excel文件：{path}（共 {len(letters)} 条记录）")
            QMessageBox.information(self, "导出成功", f"已导出至：\n{path}")
        except Exception as e:
            self._log(f"导出Excel失败：{e}")
            QMessageBox.critical(self, "导出失败", str(e))

    def _export_pdf(self):
        if not HAS_REPORTLAB:
            QMessageBox.warning(self, "缺少依赖", "未安装 reportlab 库，无法导出 PDF。\n请运行：pip install reportlab")
            self._log("导出PDF失败：缺少 reportlab 库。")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "保存PDF文件", "家书集.pdf", "PDF文件 (*.pdf)"
        )
        if not path:
            return

        try:
            font_path = self._find_chinese_font()
            if font_path:
                pdfmetrics.registerFont(TTFont("ChineseFont", font_path))
                cn_font = "ChineseFont"
            else:
                cn_font = "Helvetica"

            doc = SimpleDocTemplate(path, pagesize=A4,
                                    leftMargin=2*cm, rightMargin=2*cm,
                                    topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CNTitle", parent=styles["Title"], fontName=cn_font, fontSize=18, leading=24
            )
            heading_style = ParagraphStyle(
                "CNHeading", parent=styles["Heading2"], fontName=cn_font, fontSize=13, leading=18
            )
            body_style = ParagraphStyle(
                "CNBody", parent=styles["Normal"], fontName=cn_font, fontSize=10, leading=16
            )
            meta_style = ParagraphStyle(
                "CNMeta", parent=styles["Normal"], fontName=cn_font, fontSize=9, leading=13,
                textColor=QColor("#666").name() if False else "#666666",
            )

            story = []
            letters = self._get_filtered_handover_letters()
            filter_desc = self._get_handover_filter_desc()
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            story.append(Paragraph("家书集", title_style))
            story.append(Paragraph(f"生成时间：{now_str}　筛选条件：{filter_desc}　共 {len(letters)} 条", meta_style))
            story.append(Spacer(1, 0.5*cm))

            for lt in letters:
                title = lt.get("title") or "无标题"
                sender = lt.get("sender_name") or "未知"
                receiver = lt.get("receiver_name") or "未知"
                send_date = lt.get("send_date") or "未知"
                receive_date = lt.get("receive_date") or "未知"

                story.append(Paragraph(f"✉ {title}", heading_style))
                meta = f"寄信人：{sender}　收信人：{receiver}　寄出：{send_date}　收到：{receive_date}"
                story.append(Paragraph(meta, meta_style))
                story.append(Spacer(1, 0.3*cm))

                content = lt.get("content") or "（无内容）"
                for line in content.split("\n"):
                    story.append(Paragraph(line, body_style))

                photo_rows = execute_query_returning(
                    "SELECT file_path, description FROM photos WHERE letter_id = ?", (lt.get("id"),)
                )
                for ph in photo_rows:
                    if os.path.exists(ph["file_path"]):
                        try:
                            img = RLImage(ph["file_path"], width=10*cm, height=7*cm)
                            story.append(img)
                            if ph.get("description"):
                                story.append(Paragraph(ph["description"], meta_style))
                        except Exception:
                            story.append(Paragraph(f"[图片：{ph.get('description','')}]（加载失败）", body_style))
                    else:
                        story.append(Paragraph(f"[图片占位：{ph.get('description','') or '照片'}]", meta_style))

                story.append(Spacer(1, 0.5*cm))
                story.append(PageBreak())

            doc.build(story)
            self._log(f"已导出PDF文件：{path}（共 {len(letters)} 条记录）")
            QMessageBox.information(self, "导出成功", f"已导出至：\n{path}")
        except Exception as e:
            self._log(f"导出PDF失败：{e}")
            QMessageBox.critical(self, "导出失败", str(e))

    def _export_html(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存HTML文件", "家书集.html", "HTML文件 (*.html)"
        )
        if not path:
            return

        try:
            letters = self._get_filtered_handover_letters()
            filter_desc = self._get_handover_filter_desc()
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            html_parts = [
                "<!DOCTYPE html>",
                "<html lang='zh-CN'><head><meta charset='UTF-8'>",
                "<title>家书集</title>",
                "<style>",
                "body{font-family:'Microsoft YaHei','SimHei',sans-serif;max-width:900px;margin:0 auto;padding:20px;background:#f5f7fa;color:#333;}",
                "h1{text-align:center;color:#1a73e8;border-bottom:2px solid #1a73e8;padding-bottom:10px;}",
                ".info{text-align:center;color:#666;font-size:13px;margin-bottom:20px;}",
                ".letter{background:#fff;border-radius:8px;padding:20px;margin:20px 0;box-shadow:0 2px 8px rgba(0,0,0,0.08);}",
                ".letter h2{color:#1a73e8;margin:0 0 8px 0;font-size:18px;}",
                ".meta{color:#666;font-size:13px;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid #eee;}",
                ".content{line-height:1.8;white-space:pre-wrap;font-size:15px;}",
                ".photo{margin:10px 0;text-align:center;}",
                ".photo img{max-width:100%;border-radius:6px;box-shadow:0 2px 6px rgba(0,0,0,0.1);}",
                ".photo-desc{color:#999;font-size:12px;margin-top:4px;}",
                "</style></head><body>",
                "<h1>家书集</h1>",
                f"<div class='info'>生成时间：{now_str}　筛选条件：{filter_desc}　共 {len(letters)} 条记录</div>",
            ]

            for lt in letters:
                title = lt.get("title") or "无标题"
                sender = lt.get("sender_name") or "未知"
                receiver = lt.get("receiver_name") or "未知"
                send_date = lt.get("send_date") or "未知"
                receive_date = lt.get("receive_date") or "未知"
                content = lt.get("content") or "（无内容）"

                html_parts.append("<div class='letter'>")
                html_parts.append(f"<h2>✉ {self._html_escape(title)}</h2>")
                html_parts.append(
                    f"<div class='meta'>寄信人：{self._html_escape(sender)}　"
                    f"收信人：{self._html_escape(receiver)}　"
                    f"寄出：{self._html_escape(send_date)}　"
                    f"收到：{self._html_escape(receive_date)}</div>"
                )
                html_parts.append(f"<div class='content'>{self._html_escape(content)}</div>")

                photo_rows = execute_query_returning(
                    "SELECT file_path, description FROM photos WHERE letter_id = ?", (lt.get("id"),)
                )
                for ph in photo_rows:
                    desc = self._html_escape(ph.get("description") or "")
                    html_parts.append("<div class='photo'>")
                    if os.path.exists(ph["file_path"]):
                        html_parts.append(f"<img src='{self._html_escape(ph['file_path'])}' alt='{desc}'>")
                    else:
                        html_parts.append(f"<div style='color:#999;font-style:italic;'>[图片占位：{desc or '照片'}]</div>")
                    if desc:
                        html_parts.append(f"<div class='photo-desc'>{desc}</div>")
                    html_parts.append("</div>")

                html_parts.append("</div>")

            html_parts.append("</body></html>")

            with open(path, "w", encoding="utf-8") as f:
                f.write("\n".join(html_parts))

            self._log(f"已导出HTML文件：{path}（共 {len(letters)} 条记录）")
            QMessageBox.information(self, "导出成功", f"已导出至：\n{path}")
        except Exception as e:
            self._log(f"导出HTML失败：{e}")
            QMessageBox.critical(self, "导出失败", str(e))

    @staticmethod
    def _html_escape(text):
        return (str(text)
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace('"', "&quot;")
                .replace("'", "&#39;"))

    def _generate_handover(self):
        if not HAS_OPENPYXL:
            QMessageBox.warning(self, "缺少依赖", "未安装 openpyxl 库，无法导出清单。\n请运行：pip install openpyxl")
            self._log("生成移交清单失败：缺少 openpyxl 库。")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "保存移交清单", "移交清单.xlsx", "Excel文件 (*.xlsx)"
        )
        if not path:
            return

        try:
            rows = execute_query_returning("""
                SELECT l.id, l.title, p1.name AS sender_name, p2.name AS receiver_name,
                       l.send_date, l.receive_date, l.restoration_status, l.is_private,
                       COALESCE(br.status, '无') AS borrow_status,
                       COALESCE(br.borrower_name, '') AS borrower_name
                FROM letters l
                LEFT JOIN people p1 ON l.sender_id = p1.id
                LEFT JOIN people p2 ON l.receiver_id = p2.id
                LEFT JOIN (
                    SELECT letter_id, status, borrower_name
                    FROM borrow_records
                    WHERE id IN (SELECT MAX(id) FROM borrow_records GROUP BY letter_id)
                ) br ON br.letter_id = l.id
                ORDER BY l.id
            """)

            borrow_filter = self.filter_borrow_combo.currentText()
            if borrow_filter != "全部":
                rows = [r for r in rows if r.get("borrow_status", "无") == borrow_filter]

            restoration_filter = self.filter_restoration_combo.currentText()
            if restoration_filter != "全部":
                rows = [r for r in rows if r.get("restoration_status", "") == restoration_filter]

            if self.filter_private_check.isChecked():
                rows = [r for r in rows if r.get("is_private")]

            filter_desc = self._get_handover_filter_desc()
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "移交清单"

            info_font = XlFont(name="Microsoft YaHei", size=11, color="333333")
            ws.merge_cells("A1:I1")
            ws.cell(row=1, column=1, value="移交清单").font = XlFont(
                name="Microsoft YaHei", bold=True, size=16, color="1A73E8"
            )
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            ws.merge_cells("A2:I2")
            ws.cell(row=2, column=1, value=f"生成时间：{now_str}　筛选条件：{filter_desc}　共计 {len(rows)} 条记录").font = info_font

            headers = ["编号", "信件标题", "寄信人", "收信人", "寄出日期",
                        "收到日期", "保存状态", "借阅状态", "借阅人"]
            header_font = XlFont(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            content_font = XlFont(name="Microsoft YaHei", size=10)
            content_align = Alignment(vertical="top", wrap_text=True)
            warn_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            danger_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

            for row_idx, row in enumerate(rows, 4):
                values = [
                    row.get("id"), row.get("title", ""), row.get("sender_name", ""),
                    row.get("receiver_name", ""), row.get("send_date", ""),
                    row.get("receive_date", ""), row.get("restoration_status", ""),
                    row.get("borrow_status", "无"), row.get("borrower_name", ""),
                ]
                for col_idx, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = content_font
                    cell.alignment = content_align
                    cell.border = thin_border

                rest_status = row.get("restoration_status", "")
                if rest_status == "需修复":
                    ws.cell(row=row_idx, column=7).fill = danger_fill
                elif rest_status == "轻微损毁":
                    ws.cell(row=row_idx, column=7).fill = warn_fill

                borrow_status = row.get("borrow_status", "无")
                if borrow_status == "借出":
                    ws.cell(row=row_idx, column=8).fill = warn_fill

            col_widths = [6, 20, 10, 10, 14, 14, 10, 10, 10]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            ws.auto_filter.ref = ws.dimensions

            wb.save(path)
            self._log(f"已生成移交清单：{path}（筛选：{filter_desc}，共 {len(rows)} 条记录）")
            QMessageBox.information(self, "导出成功", f"移交清单已导出至：\n{path}")
        except Exception as e:
            self._log(f"生成移交清单失败：{e}")
            QMessageBox.critical(self, "导出失败", str(e))

    def _print_handover(self):
        try:
            rows = execute_query_returning("""
                SELECT l.id, l.title, p1.name AS sender_name, p2.name AS receiver_name,
                       l.restoration_status,
                       COALESCE(br.status, '无') AS borrow_status
                FROM letters l
                LEFT JOIN people p1 ON l.sender_id = p1.id
                LEFT JOIN people p2 ON l.receiver_id = p2.id
                LEFT JOIN (
                    SELECT letter_id, status FROM borrow_records
                    WHERE id IN (SELECT MAX(id) FROM borrow_records GROUP BY letter_id)
                ) br ON br.letter_id = l.id
                ORDER BY l.id
            """)

            html_parts = [
                "<!DOCTYPE html><html><head><meta charset='UTF-8'>",
                "<style>",
                "body{font-family:'Microsoft YaHei',sans-serif;padding:20px;}",
                "h1{text-align:center;color:#1a73e8;}",
                "table{width:100%;border-collapse:collapse;margin-top:16px;}",
                "th{background:#1a73e8;color:#fff;padding:8px;text-align:left;}",
                "td{border:1px solid #ddd;padding:6px 8px;}",
                "tr:nth-child(even){background:#f5f7fa;}",
                ".warn{color:#fbbc04;font-weight:bold;}",
                ".danger{color:#ea4335;font-weight:bold;}",
                ".good{color:#34a853;}",
                "@media print{body{padding:0;}}",
                "</style></head><body>",
                f"<h1>移交清单</h1>",
                f"<p>生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}　共计 {len(rows)} 条记录</p>",
                "<table><thead><tr>",
                "<th>编号</th><th>信件标题</th><th>寄信人</th><th>收信人</th><th>保存状态</th><th>借阅状态</th>",
                "</tr></thead><tbody>",
            ]

            for row in rows:
                rest = row.get("restoration_status", "")
                rest_cls = "good" if rest == "良好" else ("danger" if rest == "需修复" else "")
                borrow = row.get("borrow_status", "无")
                borrow_cls = "warn" if borrow == "借出" else ""
                html_parts.append(
                    f"<tr>"
                    f"<td>{row.get('id','')}</td>"
                    f"<td>{self._html_escape(row.get('title',''))}</td>"
                    f"<td>{self._html_escape(row.get('sender_name',''))}</td>"
                    f"<td>{self._html_escape(row.get('receiver_name',''))}</td>"
                    f"<td class='{rest_cls}'>{self._html_escape(rest)}</td>"
                    f"<td class='{borrow_cls}'>{self._html_escape(borrow)}</td>"
                    f"</tr>"
                )

            html_parts.append("</tbody></table></body></html>")

            tmp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_print_handover.html")
            with open(tmp_path, "w", encoding="utf-8") as f:
                f.write("\n".join(html_parts))

            from PyQt5.QtWidgets import QPlainTextEdit
            from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

            printer = QPrinter(QPrinter.HighResolution)
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                from PyQt5.QtCore import QUrl
                from PyQt5.QtWebKitWidgets import QWebPage
                try:
                    page = QWebPage()
                    page.mainFrame().load(QUrl.fromLocalFile(tmp_path))
                    page.mainFrame().evaluateJavaScript("window.print();")
                except ImportError:
                    painter = QPainter()
                    painter.begin(printer)
                    doc_font = QFont("Microsoft YaHei", 10)
                    painter.setFont(doc_font)
                    y_pos = 100
                    painter.drawText(100, y_pos, f"移交清单 - 共计 {len(rows)} 条记录")
                    y_pos += 40
                    for row in rows:
                        if y_pos > painter.device().height() - 100:
                            printer.newPage()
                            y_pos = 100
                        line = (
                            f"#{row.get('id','')}  {row.get('title','')}  "
                            f"{row.get('sender_name','')}→{row.get('receiver_name','')}  "
                            f"状态:{row.get('restoration_status','')}  借阅:{row.get('borrow_status','无')}"
                        )
                        painter.drawText(100, y_pos, line)
                        y_pos += 24
                    painter.end()

            try:
                os.remove(tmp_path)
            except OSError:
                pass

            self._log("已打印移交清单。")
        except Exception as e:
            self._log(f"打印移交清单失败：{e}")
            QMessageBox.critical(self, "打印失败", str(e))

    def _find_chinese_font(self):
        candidates = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/msyhbd.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
        ]
        for f in candidates:
            if os.path.exists(f):
                return f
        return None

    def refresh(self):
        self._refresh_backup_list()
        self._refresh_handover_table()
        self._refresh_statistics()
